"""读取 requirement.csv.xlsx 文件"""
import os

file_path = os.path.join(os.path.dirname(__file__), '..', 'DATA', 'current', 'requirement.csv.xlsx')

print(f"File exists: {os.path.exists(file_path)}")
print(f"File size: {os.path.getsize(file_path)} bytes")

try:
    import openpyxl
    wb = openpyxl.load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")
        print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
        
        # 打印所有行
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 20:  # 最多打印20行
                print(row)
except Exception as e:
    print(f"Error with openpyxl: {e}")

# 也尝试用 pandas 的不同方式
try:
    import pandas as pd
    df = pd.read_excel(file_path, header=None)
    print("\nPandas read (no header):")
    print(f"Shape: {df.shape}")
    print(df)
except Exception as e:
    print(f"Error with pandas: {e}")
